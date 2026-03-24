"""Excel utilities for GST reconciliation processing"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

from .normalization import (
    _normalize_header_token,
    rename_headers_canonically,
    unify_invoice_numbers
)


def apply_excel_formatting(output_file: str):
    """Apply formatting to the output Excel file"""
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Find remark column
    remark_col_idx = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == 'Remark':
            remark_col_idx = idx
            break
    
    # Apply red fill to mismatched rows
    if remark_col_idx is not None:
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        for row in ws.iter_rows(min_row=2, min_col=remark_col_idx, max_col=remark_col_idx):
            for cell in row:
                if cell.value == 'Mismatched':
                    for entire_row_cell in ws[cell.row]:
                        entire_row_cell.fill = red_fill
    
    # Add subtotal row
    subtotal_columns = ['Taxable Amt', 'IGST', 'CGST', 'SGST']
    subtotal_values = {}
    header = [cell.value for cell in ws[1]]
    col_letter_map = {col: ws.cell(row=1, column=i+1).column_letter for i, col in enumerate(header)}
    
    for col in subtotal_columns:
        if col in header:
            col_letter = col_letter_map[col]
            max_row = ws.max_row
            subtotal_formula = f"=SUBTOTAL(9, {col_letter}2:{col_letter}{max_row})"
            subtotal_values[col_letter] = subtotal_formula
    
    ws.insert_rows(1)
    ws['A1'] = 'SUBTOTAL'
    for col_letter, formula in subtotal_values.items():
        ws[f"{col_letter}1"] = formula
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    wb.save(output_file)


def load_and_validate_excel(file_path: str, sheet_name: str = None):
    """Load Excel file and validate required columns exist"""
    ext = os.path.splitext(file_path)[1].lower()
    if ext != ".xlsx":
        raise ValueError("Unsupported file type. Please upload a .xlsx file.")
    
    try:
        df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    except Exception as e:
        raise ValueError(f"Failed to read Excel: {e}")
    
    # Find header row
    header_row_idx = None
    for idx, row in df_raw.iterrows():
        row_values = row.fillna('').astype(str).tolist()
        norm_values = {_normalize_header_token(v) for v in row_values}
        if {"FORMS", "GSTIN NO"}.issubset(norm_values) or {"FORMS", "GSTIN"}.issubset(norm_values):
            header_row_idx = idx
            break
    
    if header_row_idx is None:
        raise ValueError("Could not find the header row. Make sure columns include 'Forms' and 'GSTIN NO.'")
    
    # Load with proper header
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row_idx)
    df.columns = df.columns.astype(str).str.strip()
    df = rename_headers_canonically(df)
    
    # Validate required columns
    required_cols = ['Forms', 'GSTIN NO.', 'Invoice No.', 'CGST', 'SGST', 'IGST']
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        raise ValueError(f"Missing required columns: {missing_cols}. Please ensure your sheet has these columns.")
    
    # Normalize invoice numbers
    df['Invoice No.'] = df['Invoice No.'].astype(str).str.strip().str.upper()
    df = unify_invoice_numbers(df)
    
    return df
