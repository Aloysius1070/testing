"""TDS calculation service - core business logic for TDS matching"""

import os
import io
import tempfile
import pandas as pd
import numpy as np
from sklearn.preprocessing import OneHotEncoder
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# No job tracking - synchronous processing


def classify_tds_entries(ledger_path):
    """
    Classify TDS entries into "Ignore" and "Ledger" categories
    from the reference file (sheet.csv)
    
    Args:
        ledger_path: Path to sheet.csv containing account classifications
    
    Returns:
        ignored_accounts: List of accounts to ignore
        ledgered_accounts: List of active ledger accounts
    """
    # Load the classification reference file
    classification_df = pd.read_csv(ledger_path)
    
    # Initialize 'Constant' column
    classification_df['Constant'] = 'Ledger'  # Default to Ledger
    
    # Mark entries to ignore
    if 'TDS Account' in classification_df.columns:
        classification_df.loc[classification_df['TDS Account'] == 'ICICI Bank', 'Constant'] = 'Ignore'
        classification_df.loc[classification_df['TDS Account'] == 'Round Off-Purchase', 'Constant'] = 'Ignore'
    
    # Get ignored and ledger accounts
    ignored_accounts = classification_df[classification_df['Constant'] == 'Ignore']['TDS Account'].dropna().tolist()
    ledgered_accounts = classification_df[classification_df['Constant'] == 'Ledger']['TDS Account'].dropna().tolist()
    
    return ignored_accounts, ledgered_accounts


def load_ledger_reference(ledger_path):
    """
    Load the ledger reference file (sheet.csv)
    
    Args:
        ledger_path: Full path to sheet.csv
    
    Returns:
        DataFrame with ledger data
    """
    if not os.path.exists(ledger_path):
        raise FileNotFoundError(f"Ledger file not found: {ledger_path}")
    
    df = pd.read_csv(ledger_path)
    return df


def process_tds_file(sheet_file_bytes, ledger_file_path):
    """
    Main TDS processing function - orchestrates the entire workflow
    
    Args:
        sheet_file_bytes: Uploaded user file content (bytes) - transaction data
        ledger_file_path: Path to sheet.csv - classification reference
    
    Returns:
        Tuple: (output_bytes, download_name)
    """
    
    try:
        # STEP 1: Load user's transaction data
        sheet_file = io.BytesIO(sheet_file_bytes)
        sheet_df = pd.read_csv(sheet_file)
        
        if 'TDS Account' not in sheet_df.columns:
            raise ValueError("Sheet must contain 'TDS Account' column")
        
        # STEP 2: Load and classify reference accounts
        ignored_accounts, ledgered_accounts = classify_tds_entries(ledger_file_path)
        
        # STEP 3: Process transaction data
        df = sheet_df.copy()
        
        # Drop ignored columns from transaction data
        df.drop(columns=ignored_accounts, errors='ignore', inplace=True)
        
        # Fill NaN values
        df.fillna(0, inplace=True)
        
        # Ensure ledger columns exist and are numeric
        existing_ledger_cols = [col for col in ledgered_accounts if col in df.columns]
        
        for col in existing_ledger_cols:
            df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # Initialize result columns
        df['Rate'] = None
        df['Value'] = None
        df['Ledger'] = None
        
        # STEP 4: TDS Matching Algorithm
        percentages = [1, 2, 5, 10, 20]  # Common TDS percentages
        total_rows = len(df)
        matched_count = 0
        
        for idx, row in df.iterrows():
            tds_val = row['TDS Account']
            
            # Skip if TDS is empty or zero
            if pd.isna(tds_val) or tds_val == 0:
                continue
            
            tds_val = float(tds_val)
            matched = False
            
            # Try each ledger column
            for col in existing_ledger_cols:
                if matched:
                    break
                
                ledger_val = row[col]
                
                # Skip if ledger value is empty or zero
                if pd.isna(ledger_val) or ledger_val == 0:
                    continue
                
                ledger_val = float(ledger_val)
                abs_val = abs(ledger_val)
                
                # Try each percentage
                for p in percentages:
                    expected_tds = abs_val * (p / 100.0)
                    
                    # Check if TDS matches expected (within tolerance)
                    if np.isclose(tds_val, expected_tds, atol=0.5):
                        df.at[idx, 'Rate'] = f"{p}%"
                        df.at[idx, 'Value'] = ledger_val
                        df.at[idx, 'Ledger'] = col
                        matched = True
                        matched_count += 1
                        break
        
        # STEP 5: Clean & filter results
        # Keep only rows with successful matches
        df.dropna(subset=['Rate'], inplace=True)
        
        # Remove ledger columns from output (they're not needed in final result)
        df.drop(columns=existing_ledger_cols, errors='ignore', inplace=True)
        
        # Reset index
        df.reset_index(drop=True, inplace=True)
        
        # STEP 6: Export to Excel
        output_file = io.BytesIO()
        
        # Create Excel writer
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='TDS Matched', index=False)
            
            # Format the Excel file
            workbook = writer.book
            worksheet = workbook['TDS Matched']
            
            # Header formatting
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = str(cell.value)
                    except:
                        pass
                
                adjusted_width = min(len(max_length) + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output_file.seek(0)
        output_bytes = output_file.getvalue()
        
        download_name = 'tds_matched_results.xlsx'
        
        print(f"TDS processing complete: {matched_count}/{total_rows} rows matched")
        
        return output_bytes, download_name
        
    except Exception as e:
        print(f"Error in process_tds_file: {str(e)}")
        raise


def validate_input_file(df):
    """Validate that uploaded file has required structure"""
    required_columns = ['TDS Account']
    
    missing = [col for col in required_columns if col not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")
    
    if len(df) == 0:
        raise ValueError("File is empty")
    
    return True
